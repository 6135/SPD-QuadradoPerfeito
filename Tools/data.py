import os.path
from os import path
import openpyxl as op
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import re
sheet = 'Hibrid'
workbook = '../relatorio/gui_data.xlsx'
directory = '../results/gui/HibridTests/'

def write_excel(results):    
    if not path.exists(workbook):
        wb = Workbook(workbook)
        
    else:
        wb = op.load_workbook(workbook)
    
    if sheet not in wb.sheetnames:
        ws = wb.create_sheet(sheet)
    else:
        ws = wb[sheet]
    
    wb.save(workbook)

    for i in range(34):#atribuir o numero de teste a cada coluna
            ws.cell(row = i+2,column = 1, value = '#'+str(i+1))

    ws.cell(row=i+1,column = 1, value = 'Média')
    ws.cell(row=i+2,column = 1, value = 'DPadrão')

    
    #distribui os dados dos ficheiros, um ficheiro por cada linha
    row = 2;
    j = 2
    type = ''
    for file in results:
        f = open(directory+file,'r')
        values = f.readlines()
        col=2   
        total = 0
        ws.cell(row = 1,column = j, value = file.partition('.')[0][1:]+"x"+file.partition('.')[0][1:]+"_Hibrid")
        j+=1
        for v in values[1:]:     
            if v != '\n' and v[0:4] == 'real':  
                #print(re.findall(r"\d*\,\d+|\d+",v)[0])  
                number = re.findall(r"\d*\.\d+|\d+",v)
                print(float(number[0])*60+float(number[1]))
                ws.cell(row = col,column = row, value = float(number[0])*60+float(number[1])*1000)
                col+=1
                
                total += float(number[0])*60+float(number[1])
        
        ws.cell(row=col,column=row,value=(total/(len(values)-1))*pow(1,-3)*1000)
        type = file[0:1]
        row+=1    

    wb.save(workbook)
    wb.close()



def main():
    results = os.listdir(directory)
    sort_by_type = sorted(results, key=lambda x:  (x[0:1],int(x[1:].partition('.')[0])))
    #for item in sort_by_type:
    #    print(item)
    write_excel(sort_by_type)

if __name__ == "__main__":
	main()
